# -*- coding: utf-8 -*-
import os
import pandas as pd
import glob
import difflib
import json
import time
import copy
try:
    from Emb_data import run_embedding
    from DBS_data import run_clustering
    from kmeans_data import run_kmeans_partition
    try:
        from DBreview_data import run_review
        HAS_REVIEW_MODULE = True
    except ImportError:
        HAS_REVIEW_MODULE = False
        print("警告: DBreview_data 未找到，将跳过聚类复查步骤。")
    HAS_NEW_MODULES = True
except ImportError as e:
    HAS_NEW_MODULES = False
    print(f"警告: Emb_data 或 DBS_data 或 kmeans_data 未找到 ({e})，将跳过高级聚类步骤。")

# 兼容 Python 2/3
try:
    basestring
except NameError:
    basestring = str

def _script_dir():
    """获取脚本所在目录"""
    return os.path.dirname(os.path.abspath(__file__))

def load_config():
    """加载配置文件"""
    cfg_path = os.path.join(_script_dir(), "json", "config.json")
    if os.path.isfile(cfg_path):
        with open(cfg_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def find_latest_file(directory):
    """查找目录下最新的 .xlsx 文件"""
    search_pattern = os.path.join(directory, "*.xlsx")
    files = glob.glob(search_pattern)
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return None
    files.sort(key=os.path.getmtime)
    return files[-1]

def is_similar(s1, s2, threshold=0.75):
    """
    计算两个字符串的相似度
    :param s1: 字符串1
    :param s2: 字符串2
    :param threshold: 相似度阈值
    :return: Boolean (是否相似)
    """
    if not isinstance(s1, basestring): s1 = str(s1)
    if not isinstance(s2, basestring): s2 = str(s2)
    return difflib.SequenceMatcher(None, s1, s2).ratio() > threshold

def semantic_deduplicate(df, threshold=0.75, is_test=False):
    """
    语义去重 (基于 Difflib)
    :param df: 数据 DataFrame
    :param threshold: 相似度阈值
    :return: (kept_df, removed_count, logs)
    """
    if df.empty:
        return df, 0, []
        
    subjects = df["subject"].tolist()
    indices = df.index.tolist()
    
    keep_indices = []
    kept_subjects = []
    removed_logs = []
    
    start_time = time.time()
    
    # 尝试引入 tqdm
    try:
        from tqdm import tqdm
        iter_subjects = tqdm(enumerate(subjects), total=len(subjects), desc="Difflib 语义去重")
    except ImportError:
        iter_subjects = enumerate(subjects)
        print(f"正在进行语义去重 (阈值: {threshold}), 共 {len(subjects)} 条...")
    
    for i, subj in iter_subjects:
        is_dup = False
        # 与已保留的条目逐一对比 (O(N^2) 复杂度，量大时较慢)
        for kept_subj in kept_subjects:
            if is_similar(subj, kept_subj, threshold):
                is_dup = True
                if is_test:
                    removed_logs.append({
                        "stage": "semantic_dedup",
                        "removed": subj,
                        "kept": kept_subj,
                        "similarity": f"Difflib > {threshold}"
                    })
                break
        
        if not is_dup:
            keep_indices.append(indices[i])
            kept_subjects.append(subj)
            
    end_time = time.time()
    print(f"语义去重耗时: {end_time - start_time:.2f}秒")
    return df.loc[keep_indices], len(df) - len(keep_indices), removed_logs

def read_excel_fallback(path):
    """读取 Excel 的回退方法 (Openpyxl)"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        data = []
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                headers.append(cell.value)
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(row)
        except TypeError:
            for row in ws.iter_rows(min_row=2):
                data.append([cell.value for cell in row])
        df = pd.DataFrame(data, columns=headers)
        return df
    except Exception as e:
        print(f"Openpyxl 读取失败: {e}")
        raise e

def run_deduplication(input_data=None, extra_sheets=None):
    """
    去重流程主入口
    流程: 简单去重 -> 语义去重 -> 向量化 -> 聚类
    """
    script_dir = _script_dir()
    df = None
    
    # 1. 加载数据
    if isinstance(input_data, pd.DataFrame):
        df = input_data.copy()
        # 删除不需要的列
        cols_to_drop = ["time", "source_type"]
        df = df.drop(columns=cols_to_drop, errors="ignore")
        print(f"\n[去重阶段] 接收内存数据: {len(df)} 条")
    else:
        # 从文件加载
        target_file = input_data
        if not target_file:
            cleandata_dir = os.path.join(script_dir, "cleandata")
            if not os.path.isdir(cleandata_dir):
                return {"success": False, "message": "未找到 cleandata 文件夹"}
            target_file = find_latest_file(cleandata_dir)
            if not target_file:
                return {"success": False, "message": "无可用数据文件"}
        
        print(f"\n[去重阶段] 处理文件: {target_file}")
        try:
            df = pd.read_excel(target_file)
        except:
            try:
                df = read_excel_fallback(target_file)
            except Exception as e:
                return {"success": False, "message": f"读取失败: {e}"}

    if "subject" not in df.columns:
        return {"success": False, "message": "缺少 'subject' 列"}

    # 采样逻辑 (仅用于调试交互)
    # try:
    #     user_input = input("请输入采样数量 (回车跳过): ").strip()
    #     if user_input:
    #         sample_size = int(user_input)
    #         if sample_size < len(df):
    #             df = df.sample(n=sample_size)
    #             print(f"已采样 {sample_size} 条数据。")
    # except: pass

    cfg = load_config()
    is_test = cfg.get("isTest", False)
    thresholds = cfg.get("deduplication_thresholds", {"simple_dedup": "first", "semantic_dedup": 0.75})
    
    original_count = len(df)
    
    try:
        # 预处理：确保 subject 是字符串并去除首尾空格，避免因空格导致的去重失效
        if 'subject' in df.columns:
            df['subject'] = df['subject'].astype(str).str.strip()

        # Step 1: 简单去重 (完全匹配)
        df_simple = df.drop_duplicates(subset=["subject"], keep="first")
        print(f"简单去重: 移除 {original_count - len(df_simple)} 条重复数据。")

        # Step 2: 语义去重 (Difflib)
        semantic_threshold = thresholds.get("semantic_dedup", 0.75)
        df_semantic, semantic_removed_count, semantic_logs = semantic_deduplicate(
            df_simple, threshold=semantic_threshold, is_test=is_test
        )
        print(f"语义去重: 移除 {semantic_removed_count} 条相似数据。")

        # Step 3: 高级聚类 (Embedding + Clustering)
        cluster_logs = []
        df_final = df_semantic
        
        # 如果没有安装 Embedding 模块，或者配置关闭了聚类，则跳过
        # 但通常我们希望尽可能聚类
        if HAS_NEW_MODULES:
            print("\n>>> 子步骤: 向量化与聚类 <<<")
            # 3.1 向量化
            df_emb, emb_path = run_embedding(df_semantic, column='subject')
            
            # 3.2 聚类
            if emb_path:
                # 加载 K-Means 配置
                kmeans_cfg = cfg.get("kmeans_settings", {})
                use_kmeans = kmeans_cfg.get("enabled", True)
                
                if use_kmeans:
                    # 自动计算 n_clusters (如果配置了 target_batch_size)
                    n_clusters = kmeans_cfg.get("n_clusters", 10)
                    target_batch = kmeans_cfg.get("target_batch_size", 70)
                    
                    # 如果总数较大，根据 target_batch_size 调整 n_clusters
                    total_samples = len(df_emb)
                    if target_batch > 0 and total_samples > target_batch:
                        calc_n = total_samples // target_batch
                        if calc_n > 1:
                            n_clusters = calc_n
                            # print(f"根据目标批次大小 {target_batch}，调整 n_clusters = {n_clusters}")
                    
                    # 执行 K-Means 分区
                    batches = run_kmeans_partition(df_emb, emb_path, n_clusters=n_clusters)
                    
                    # 分批执行 HDBSCAN
                    all_clustered_dfs = []
                    max_cluster_id_offset = 0
                    
                    # 进度条
                    try:
                        from tqdm import tqdm
                        iter_batches = tqdm(batches, desc="HDBSCAN 分批聚类进度")
                    except ImportError:
                        iter_batches = batches
                        print(f"开始分批聚类处理，共 {len(batches)} 个批次...")
                    
                    for batch_id, b_df, b_emb in iter_batches:
                        # 将 b_emb 转为 list 存入 b_df (run_clustering 需要)
                        # 或者直接传给 run_clustering (如果它支持)
                        # 现在的 run_clustering 支持 df 和 emb_path，但我们已经切分了 embeddings，没法传 path
                        # 需要稍微修改 run_clustering 或者将 embeddings 放入 df
                        
                        # 这里的 b_df 应该已经包含 embedding_index，但我们现在有直接的 b_emb (numpy array)
                        # 为了兼容，我们将 b_emb 挂载到 b_df 上，或者修改 run_clustering 接受 embeddings 参数
                        # 最简单的: 临时修改 b_df['embedding'] = list(b_emb)
                        # 注意: b_emb 是 numpy array
                        b_df_temp = b_df.copy()
                        b_df_temp['embedding'] = list(b_emb)
                        
                        # 运行 HDBSCAN
                        # 注意: 这里 emb_path 传 None，因为我们已经在 df 里放了 embedding
                        b_clustered, b_logs = run_clustering(b_df_temp, emb_path=None)
                        
                        # 处理 Cluster ID 偏移
                        # 只有有效的 cluster_id (>=0) 才需要偏移
                        # 噪点 -1 保持 -1
                        if not b_clustered.empty and 'cluster_id' in b_clustered.columns:
                            mask_valid = b_clustered['cluster_id'] != -1
                            if mask_valid.any():
                                # 当前批次的最大 ID
                                current_max = b_clustered.loc[mask_valid, 'cluster_id'].max()
                                
                                # 应用偏移
                                b_clustered.loc[mask_valid, 'cluster_id'] += max_cluster_id_offset
                                
                                # 更新日志中的 ID
                                for log in b_logs:
                                    if log['Cluster ID'] != -1:
                                        log['Cluster ID'] += max_cluster_id_offset
                                
                                # 更新 offset
                                # 下一批次的 ID 从 (当前最大 + 1) 开始
                                # 注意: 如果 current_max 是 NaN (空批次)，则不加
                                if not pd.isna(current_max):
                                    max_cluster_id_offset += (current_max + 1)
                        
                        # 聚类复查 (针对该批次)
                        if HAS_REVIEW_MODULE:
                            # 尝试保存临时向量文件供 run_review 使用
                            try:
                                import numpy as np # 确保 numpy 被导入
                                # 将临时文件存放到 cache 目录，避免污染根目录
                                cache_dir = os.path.join(_script_dir(), "cache")
                                if not os.path.exists(cache_dir):
                                    os.makedirs(cache_dir, exist_ok=True)
                                    
                                tmp_emb_path = os.path.join(cache_dir, f"temp_emb_batch_{batch_id}.npy")
                                np.save(tmp_emb_path, b_emb)
                                b_clustered, review_logs = run_review(b_clustered, tmp_emb_path)
                                if review_logs:
                                    # 同样需要更新 ID
                                    for log in review_logs:
                                        if 'old_cluster_id' in log: log['old_cluster_id'] += max_cluster_id_offset
                                        pass
                                    b_logs.extend(review_logs)
                                
                                if os.path.exists(tmp_emb_path):
                                    os.remove(tmp_emb_path)
                            except Exception as e:
                                print(f"批次 {batch_id} 复查失败: {e}")
    
                        all_clustered_dfs.append(b_clustered)
                        cluster_logs.extend(b_logs)
                    
                    # 合并所有批次结果
                    if all_clustered_dfs:
                        df_final = pd.concat(all_clustered_dfs, ignore_index=True)
                    else:
                        df_final = df_emb.copy()
                        df_final['cluster_id'] = -1
                        
                else:
                    # 不启用 K-Means，走原有流程
                    df_clustered, cluster_logs = run_clustering(df_emb, emb_path=emb_path)
                    df_final = df_clustered
                    
                    # 3.3 聚类复查与优化
                    review_logs = []
                    if HAS_REVIEW_MODULE:
                        df_final, review_logs = run_review(df_final, emb_path)
                        if review_logs:
                             # 合并日志以便输出
                             cluster_logs.extend(review_logs)
            else:
                print("向量化未生成文件，跳过聚类。")
        else:
            # 如果没有聚类模块，必须手动添加 cluster_id 列，否则后续 AI 分析会报错
            # cluster_id = -1 表示噪点/未聚类
            print("警告: 缺失聚类模块，所有数据标记为未聚类(-1)。")
            df_final['cluster_id'] = -1
            df_final['cohesion'] = 0.0
        
        final_count = len(df_final)
        print(f"去重流程结束: 剩余 {final_count} 条。")

        # Step 4: 保存结果
        timestamp = time.strftime("%Y%m%d")
        cleandata_dir = os.path.join(script_dir, "cleandata")
        if not os.path.isdir(cleandata_dir):
            os.makedirs(cleandata_dir, exist_ok=True)
            
        existing_files = [f for f in os.listdir(cleandata_dir) if f"deduplicated_data_{timestamp}" in f]
        version = len(existing_files) + 1
        output_filename = f"deduplicated_data_{timestamp}_v{version}.xlsx"
        output_path = os.path.join(cleandata_dir, output_filename)
        
        # 清理中间列
        # 注意: 保留 task_id 以便追溯
        # 保留 needs_image 和 needs_image_reason
        cols_to_drop = ['Cluster', 'Core Topic', 'embedding', 'embedding_index', 'normalized_subject']
        df_final_clean = df_final.drop(columns=cols_to_drop, errors='ignore')
        
        # 调整列顺序: task_id 最前，image 后面紧跟 needs_image
        cols = list(df_final_clean.columns)
        new_cols = []
        if 'task_id' in cols:
            new_cols.append('task_id')
        
        # 其他列
        for c in cols:
            if c != 'task_id':
                new_cols.append(c)
                
        # 调整 needs_image 到 image 后面
        if 'image' in new_cols and 'needs_image' in new_cols:
            idx_img = new_cols.index('image')
            # 先移除 needs_image
            new_cols.remove('needs_image')
            # 插入到 image 后面
            new_cols.insert(idx_img + 1, 'needs_image')
            
            # reason 也紧跟
            if 'needs_image_reason' in new_cols:
                new_cols.remove('needs_image_reason')
                new_cols.insert(idx_img + 2, 'needs_image_reason')
                
        df_final_clean = df_final_clean[new_cols]
        
        return {
            "success": True,
            "original_count": original_count,
            "dedup_count": final_count,
            "output_path": "Memory DataFrame (No File Saved)",
            "df_final": df_final_clean,
            "extra_logs": {
                "cluster_logs": cluster_logs,
                "extra_sheets": extra_sheets
            }
        }

    except Exception as e:
        print(f"去重过程错误: {e}")
        return {"success": False, "message": str(e)}

if __name__ == "__main__":
    run_deduplication()
