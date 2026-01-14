# -*- coding: utf-8 -*-
import os
import pandas as pd
import glob
import json
import random
import time
import argparse
from llm_client import LLMClient
import concurrent.futures
import re
import sys

def _script_dir():
    return os.path.dirname(os.path.abspath(__file__))

def find_latest_file(directory):
    """查找目录下最新的xlsx文件"""
    search_pattern = os.path.join(directory, "*.xlsx")
    files = glob.glob(search_pattern)
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return None
    files.sort(key=os.path.getmtime)
    return files[-1]

def read_excel_fallback(path):
    """Fallback to read excel using openpyxl directly"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        data = []
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                headers.append(cell.value)
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(row)
        except TypeError:
            for row in ws.iter_rows(min_row=2):
                data.append([cell.value for cell in row])
        df = pd.DataFrame(data, columns=headers)
        return df
    except Exception as e:
        print("Openpyxl fallback failed: {}".format(e))
        return pd.DataFrame()

def load_data():
    script_dir = _script_dir()
    cleandata_dir = os.path.join(script_dir, "cleandata")
    
    # Try to find the latest deduplicated data first, if not, try raw data
    target_file = find_latest_file(cleandata_dir)
    if not target_file:
        rawdata_dir = os.path.join(script_dir, "rawdata")
        target_file = find_latest_file(rawdata_dir)
        
    if not target_file:
        print("Error: No data file found in cleandata or rawdata.")
        return None
        
    print(f"Loading data from: {target_file}")
    
    try:
        df = pd.read_excel(target_file, engine='openpyxl')
    except:
        try:
            df = pd.read_excel(target_file)
        except:
            df = read_excel_fallback(target_file)
            
    if "subject" not in df.columns:
        # Try to find a column that looks like subject
        for col in df.columns:
            if "主题" in str(col) or "描述" in str(col) or "内容" in str(col):
                print(f"Using column '{col}' as subject.")
                df.rename(columns={col: "subject"}, inplace=True)
                break
        
        if "subject" not in df.columns:
            print("Error: Could not find 'subject' column.")
            return None
            
    return df

def analyze_batch(client, items):
    prompt = """
请判断以下Bug描述是否需要查看截图/图片来辅助定位。

**判断标准（非常严格，克制看图）：**
1. **必须看图**：描述中没有任何具体信息（如"如图"、"看图"、"这啥"），完全依赖图片。
2. **必须看图**：描述有Bug信息，但缺乏关键定位信息（如只说"太卡了"、"模型坏了"、"这里有Bug"），且没有提到具体的地图、NPC、道具或时间点。
3. **可能看图**：字数极少（<10字）且信息模糊。
4. **不需要看图**：描述清晰，包含具体位置、道具名、NPC名或详细的Bug现象。
5. **不需要看图**：字数>10字，且包含一定信息量。

**Input List:**
"""
    for item in items:
        prompt += f"ID_{item['id']}: {item['subject']}\n"
        
    prompt += """
**Output Format (JSON only):**
{
  "ID_x": {"needs_image": true/false, "reason": "详细理由"},
  ...
}
"""
    try:
        response = client.chat_completion([{"role": "user", "content": prompt}], json_mode=True)
        if response:
            clean_json = response.replace("```json", "").replace("```", "").strip()
            return json.loads(clean_json)
    except Exception as e:
        print(f"Batch analysis failed: {e}")
        return {}
    return {}

def find_image_column(df):
    cols = list(df.columns)
    for col in cols:
        name = str(col).lower()
        if ("image" in name) or ("img" in name) or ("图片" in str(col)) or ("截图" in str(col)) or ("pic" in name) or ("photo" in name) or ("链接" in str(col)) or ("url" in name):
            return col
    return None

def fallback_judge(text):
    s = str(text or "").strip()
    if not s:
        return True, "描述为空或仅包含空白，需看图"
    content = re.sub(r"[\s\W_]+", "", s)
    length = len(content)
    vague = re.search(r"(卡|卡顿|慢|不行|不对|有问题|坏了|闪退|崩|卡死|bug|异常)", s.lower())
    has_location = re.search(r"(地图|副本|npc|道具|装备|技能|任务|按钮|界面|设置|分辨率|帧率|服务器|时间|坐标|位置|名字)", s.lower())
    if length <= 10 and (not has_location):
        return True, "字数极少且信息模糊/缺定位信息，需看图"
    return False, "描述较清晰或信息量充足，不需要看图"

def get_result_for_id(results, id_val):
    key1 = f"ID_{id_val}"
    if key1 in results:
        return results[key1]
    for k, v in results.items():
        ks = str(k).replace(" ", "")
        if ks == key1 or str(k) == str(id_val):
            return v
    return None

def main():
    print("=== Image Check Tool ===")
    
    # 1. Load Data
    df = load_data()
    if df is None or df.empty:
        return

    total_rows = len(df)
    print(f"Total records loaded: {total_rows}")
    image_col = find_image_column(df)
    
    # 2. User Input for Sampling
    try:
        user_input = input("请输入要分析的条目数 (直接回车分析所有, 输入数字则随机抽取): ").strip()
        if user_input:
            sample_size = int(user_input)
            if sample_size < total_rows:
                df = df.sample(n=sample_size)
                print(f"已随机抽取 {sample_size} 条数据进行分析。")
            else:
                print(f"输入数量大于总数，将分析所有 {total_rows} 条数据。")
    except ValueError:
        print("输入无效，将分析所有数据。")

    # 3. Initialize LLM Client
    try:
        client = LLMClient()
    except Exception as e:
        print(f"Failed to initialize LLM Client: {e}")
        return

    # 4. Process in Batches
    items_to_process = []
    for idx, row in df.iterrows():
        items_to_process.append({"id": idx, "subject": str(row["subject"])})
        
    batch_size = 10
    batches = [items_to_process[i:i + batch_size] for i in range(0, len(items_to_process), batch_size)]
    
    results = {}
    
    print(f"开始分析 {len(items_to_process)} 条数据...")
    start_time = time.time()
    
    total_batches = len(batches)
    completed_batches = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        future_to_batch = {executor.submit(analyze_batch, client, batch): batch for batch in batches}
        print(f"分析进度: 0/{total_batches} 批次 (0.0%)")
        for future in concurrent.futures.as_completed(future_to_batch):
            batch_res = future.result()
            results.update(batch_res)
            completed_batches += 1
            percent = (completed_batches / float(total_batches)) * 100 if total_batches else 100.0
            print(f"\r分析进度: {completed_batches}/{total_batches} 批次 ({percent:.1f}%)", end="", flush=True)
        print()
            
    # 5. Output Results
    print("\n=== 分析结果 ===")
    needs_image_count = 0
    
    output_data = []
    
    for item in items_to_process:
        res = get_result_for_id(results, item['id'])
        if res is None:
            needs_image, reason = fallback_judge(item['subject'])
        else:
            needs_image = res.get("needs_image", False)
            reason = res.get("reason") or "模型输出理由缺失，已按规则补全"
        image_val = ""
        if needs_image and image_col and image_col in df.columns:
            try:
                val = df.loc[item['id'], image_col]
                if not pd.isna(val):
                    image_val = str(val)
            except Exception:
                image_val = ""
        
        output_data.append({
            "Original ID": item['id'],
            "Subject": item['subject'],
            "Needs Image": "Yes" if needs_image else "No",
            "Image": image_val,
            "Reason": reason
        })
        
        if needs_image:
            needs_image_count += 1
            
    # Convert to DataFrame and print/save
    result_df = pd.DataFrame(output_data)
    
    print(f"分析完成，耗时 {time.time() - start_time:.2f}s")
    print(f"需要看图的条目数: {needs_image_count} / {len(items_to_process)}")
    print(f"看图比例: {needs_image_count / len(items_to_process):.2%}")
    
    # Print sample of needs image
    print("\n[需要看图的典型样本]")
    print(result_df[result_df["Needs Image"] == "Yes"].head(10).to_string(index=False))
    
    readimage_dir = os.path.join(_script_dir(), "readimage")
    if not os.path.isdir(readimage_dir):
        os.makedirs(readimage_dir)
    pattern = os.path.join(readimage_dir, "image_check_result_v*.xlsx")
    files = glob.glob(pattern)
    max_v = 0
    for f in files:
        base = os.path.basename(f)
        m = re.search(r"_v(\d+)\.xlsx$", base)
        if m:
            v = int(m.group(1))
            if v > max_v:
                max_v = v
    next_v = max_v + 1
    output_file = os.path.join(readimage_dir, f"image_check_result_v{next_v}.xlsx")
    result_df.to_excel(output_file, index=False)
    print(f"\n完整结果已保存至: {output_file}")

if __name__ == "__main__":
    main()
